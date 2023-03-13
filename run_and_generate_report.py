#!/usr/bin/env python3

import argparse
import logging
import os
import sys
import subprocess
import glob
import openpyxl

#setting logger
logger = logging.getLogger('run_and_generate_report')
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

stream_handler = logging.StreamHandler()
stream_handler.setFormatter(formatter)
logger.addHandler(stream_handler)


def parse_args():
    parser = argparse.ArgumentParser(description='run_and_generate_report')
    parser.add_argument('--dataset_path', type=str, required=True, help='Path to KITTI dataset')
    parser.add_argument('--output_dir', type=str, required=True, help='Path to output directory')
    args = parser.parse_args()
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)
    args = parser.parse_args()
    return args


def create_folder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        logger.error('Creating directory. ' +  directory)
        sys.exit(1)


def run_rtapmap_all_features(output_dir, dataset_path):
    for index in range(16):
        logger.info('{}th feature start!'.format(index))
        create_folder(output_dir + str(index))
        try:
            output_text = subprocess.check_output([
                '/root/rtabmap_install/bin/rtabmap-kitti_dataset',
                '--Rtabmap/PublishRAMUsage', 'true',
                '--Rtabmap/DetectionRate', '2',
                '--Rtabmap/CreateIntermediateNodes', 'true',
                '--RGBD/LinearUpdate', '0',
                '--GFTT/QualityLevel', '0.01',
                '--GFTT/MinDistance', '7',
                '--SuperPoint/ModelPath', dataset_path+'superpoint.pt',
                '--OdomF2M/MaxSize', '3000',
                '--Mem/STMSize', '30',
                '--Kp/MaxFeatures', '750',
                '--Kp/DetectorStrategy', str(index),
                '--Vis/MaxFeatures', '1500',
                '--Vis/FeatureType', str(index),
                '--output', output_dir+str(index),
                '--gt', dataset_path+'data_odometry_poses/poses/07.txt',
                dataset_path+'data_odometry_gray/dataset/sequences/07'],
                encoding='utf-8')
        except subprocess.CalledProcessError:
            logger.error('Fail to excute {}th feature.'.format(index))
            sys.exit(1)
        logger.info('{}th feature finish!'.format(index))
        logger.info('Trying to save output text')
        with open(output_dir+'{}.txt'.format(index), 'w') as file:
            file.write(output_text)
        logger.info('Save complete!')


def parse_time_error_data(output_dir):
    logger.info('Parsing time data')
    time_dataset = {}
    error_dataset = {}
    feature_name = ['SURF', 'SIFT', 'ORB', 'FAST/FREAK',
                    'FAST/BRIEF', 'GFTT/FREAK', 'GFTT/BRIEF', 'BRISK',
                    'GFTT/ORB', 'KAZE', 'ORB-OCTREE', 'SuperPoint',
                    'SURF/FREAK', 'GFTT/DAISY', 'SURF/DAISY', 'PyDetector']
    file_path_list = glob.glob(output_dir+'*.txt')
    for file_path in file_path_list:
        index = int(file_path.split('.')[0].split('/')[-1])
        time_data = {'camera':{}, 'odom':{}, 'slam':{}}
        error_dataset[feature_name[index]] = {}

        with open(file_path, 'r') as file:
            lines = file.readlines()

        camera_time = []
        odom_time = []
        slam_time = []
        for line in lines:
            words = line.split()
            if not words:
                continue
            if words[0][:-1] == 'translational_rmse':
                error_dataset[feature_name[index]]['translational_rmse'] = float(words[-2])
            elif words[0][:-1] == 'rotational_rmse':
                error_dataset[feature_name[index]]['rotational_rmse'] = float(words[-2])
            elif words[0] == 'Iteration':
                camera_time.append(int(words[3].split('=')[-1][:-3]))
                odom_time.append(int(words[5].split('=')[-1][:-3]))
                slam_time.append(int(words[6].split('=')[-1][:-3]))
        try:
            time_data['camera']['average'] = int(sum(camera_time) / len(camera_time))
            time_data['camera']['median'] = sorted(camera_time)[len(camera_time)//2]
            time_data['camera']['max'] = max(camera_time)
            time_data['camera']['min'] = min(camera_time)

            time_data['odom']['average'] = int(sum(odom_time) / len(odom_time))
            time_data['odom']['median'] = sorted(odom_time)[len(odom_time)//2]
            time_data['odom']['max'] = max(odom_time)
            time_data['odom']['min'] = min(odom_time)

            time_data['slam']['average'] = int(sum(slam_time) / len(slam_time))
            time_data['slam']['median'] = sorted(slam_time)[len(slam_time)//2]
            time_data['slam']['max'] = max(slam_time)
            time_data['slam']['min'] = min(slam_time)
        except ZeroDivisionError:
            logger.error('Cannot divide into zero! Check gt path!')
            sys.exit(1)
        except IndexError:
            logger.error('Cannot access to wrong index!')
            sys.exit(1)

        if not camera_time or not odom_time or not slam_time:
            logger.warning('Cannot parse time data from ouput! Check command!')
        time_dataset[feature_name[index]] = time_data
    logger.info('Parsing done')
    return time_dataset, error_dataset


def write_excel(time_dataset, error_dataset, output_dir):
    logger.info('Trying to write data in excel format')
    workbook = openpyxl.load_workbook('/root/slam_team2/data_format.xlsx', data_only=True)
    worksheet = workbook.active
    #write time dataset in a row
    for row, (name, time_data) in enumerate(time_dataset.items()):
        for type, data in time_data.items():
            if type == 'camera':
                index = 0
            elif type == 'odom':
                index =17
            elif type == 'slam':
                index = 34
            worksheet.cell(row+index+2, 2, name)
            worksheet.cell(row+index+2, 1, type)
            for i, value in enumerate(data.values()):
                worksheet.cell(row+index+2, 3+i, value)
    logger.info('Writing time data done!')
    #write error dataset in a row
    for row, (name, error_data) in enumerate(error_dataset.items()):
        worksheet.cell(row+2, 8, name)
        for type, data in error_data.items():
            if type == 'translational_rmse':
                worksheet.cell(row+2, 9, data)
            elif type == 'rotational_rmse':
                worksheet.cell(row+2, 10, data)
    logger.info('Writing error data done!')
    create_folder(output_dir+'report')
    workbook.save(output_dir+'report/result.xlsx')
    logger.info('Saving '+output_dir+'report/result.xlsx complete!')


if __name__ == '__main__':
    args = parse_args()
    create_folder(args.output_dir)
    run_rtapmap_all_features(args.output_dir, args.dataset_path)
    time_dataset, error_dataset = parse_time_error_data(args.output_dir)
    write_excel(time_dataset, error_dataset, args.output_dir)
